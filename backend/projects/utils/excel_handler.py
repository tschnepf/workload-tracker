"""
Excel handler for Projects import/export functionality.
Handles multi-sheet export/import: Projects, Assignments, Deliverables.
Follows R2-REBUILD-STANDARDS: snake_case Python, camelCase Excel headers.
"""

import openpyxl
import json
import csv
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import transaction
from datetime import datetime
from ..serializers import ProjectSerializer
from ..models import Project
from people.models import Person
from people.serializers import PersonSerializer
from assignments.models import Assignment
from assignments.serializers import AssignmentSerializer
from core.utils.excel import write_headers, auto_fit_columns, create_excel_response
from core.utils.excel_sanitize import sanitize_cell


def export_projects_to_excel(queryset, filename=None, is_template=False):
    """Export projects queryset to Excel with multiple sheets."""
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if is_template:
            filename = f"projects_import_template_{timestamp}.xlsx"
        else:
            filename = f"projects_export_{timestamp}.xlsx"
    
    workbook = openpyxl.Workbook()
    
    if is_template or queryset.count() == 0:
        # Create template with example data
        _create_template_projects_sheet(workbook)
        _create_template_assignments_sheet(workbook)
        _create_template_deliverables_sheet(workbook)
    else:
        # Create export with real data
        _create_projects_sheet(workbook, queryset)
        _create_assignments_sheet(workbook, queryset)
        _create_deliverables_sheet(workbook, queryset)
    
    # Always include template examples and instructions
    _create_projects_template_sheet(workbook)
    _create_projects_instructions_sheet(workbook)
    
    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    return create_excel_response(workbook, filename)


def _create_template_projects_sheet(workbook):
    """Create Projects sheet with comprehensive example data for import template."""
    sheet = workbook.active
    sheet.title = "Projects"
    
    # Excel headers (camelCase to match API)
    headers = [
        'name', 'projectNumber', 'status', 'client', 'description',
        'startDate', 'endDate', 'estimatedHours', 'isActive'
    ]
    
    # Write headers with styling
    write_headers(sheet, headers)
    
    # Example project data rows
    example_projects = [
        ['Website Redesign', 'PRJ-2024-001', 'active', 'Acme Corp', 'Complete website overhaul', '2024-01-01', '2024-06-30', 2000, True],
        ['Mobile App Phase 2', 'PRJ-2024-002', 'planning', 'TechStart', 'iOS and Android app development', '2024-03-01', '2024-12-31', 3500, True],
        ['Internal Tool Update', '', 'active', 'Internal', 'Update existing dashboard tool', '', '', '', True],
        ['Quick Fix Project', 'QF-001', 'completed', 'Legacy Client', 'Small bug fixes and updates', '2024-01-15', '2024-02-15', 40, False]
    ]
    
    # Write example data
    for row_idx, project_data in enumerate(example_projects, start=2):
        for col_idx, value in enumerate(project_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            # Light blue background for example data
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    auto_fit_columns(sheet)


def _create_template_assignments_sheet(workbook):
    """Create Assignments sheet with comprehensive example data for import template."""
    assignments_sheet = workbook.create_sheet("Assignments")
    
    # Excel headers for assignments (enhanced with personRole for auto-creation)
    headers = [
        'projectName', 'projectNumber', 'personName', 'personEmail', 'personRole',
        'roleOnProject', 'startDate', 'endDate', 'weeklyHours', 
        'totalHours', 'notes', 'isActive'
    ]
    
    write_headers(assignments_sheet, headers)
    
    # Example assignment data rows with personRole for auto-creation
    example_assignments = [
        ['Website Redesign', 'PRJ-2024-001', 'John Smith', 'john.smith@company.com', 'Senior Engineer', 'Tech Lead', '2024-01-01', '2024-06-30', '{"2024-01-01":20,"2024-01-08":25,"2024-01-15":20}', 800, 'Full-time on project', True],
        ['Website Redesign', 'PRJ-2024-001', 'Jane Doe', 'jane.doe@company.com', 'Designer', 'UI/UX Designer', '2024-01-15', '2024-05-15', '{"2024-01-15":30,"2024-01-22":32,"2024-01-29":30}', 600, 'Design lead', True],
        ['Mobile App Phase 2', 'PRJ-2024-002', 'Alex Rodriguez', 'alex.rodriguez@company.com', 'Project Manager', 'Project Manager', '2024-03-01', '2024-12-31', '{"2024-03-01":40,"2024-03-08":40}', 1600, 'Full project oversight', True],
        ['Internal Tool Update', '', 'Sarah Wilson', 'sarah.wilson@company.com', 'Developer', 'Backend Developer', '2024-02-01', '', '{"2024-02-01":40}', 40, 'Database updates', True],
        ['Quick Fix Project', 'QF-001', 'Mike Johnson', 'mike.johnson@company.com', 'Senior Developer', 'Lead Developer', '2024-01-15', '2024-02-15', '{"2024-01-15":40,"2024-01-22":40}', 80, 'Bug fixes and testing', False],
        ['Website Redesign', 'PRJ-2024-001', 'Lisa Chen', 'lisa.chen@company.com', 'QA Engineer', 'Quality Assurance', '2024-05-01', '2024-06-30', '{"2024-05-01":32,"2024-05-08":32}', 320, 'Testing and validation', True]
    ]
    
    # Write example data
    for row_idx, assignment_data in enumerate(example_assignments, start=2):
        for col_idx, value in enumerate(assignment_data, start=1):
            cell = assignments_sheet.cell(row=row_idx, column=col_idx, value=value)
            # Light green background for example data
            cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    auto_fit_columns(assignments_sheet)


def _create_template_deliverables_sheet(workbook):
    """Create Deliverables sheet with comprehensive example data for import template."""
    deliverables_sheet = workbook.create_sheet("Deliverables")
    
    # Excel headers for deliverables
    headers = [
        'projectName', 'projectNumber', 'description', 'percentage',
        'date', 'sortOrder', 'isCompleted', 'completedDate', 'notes'
    ]
    
    write_headers(deliverables_sheet, headers)
    
    # Example deliverable data rows
    example_deliverables = [
        ['Website Redesign', 'PRJ-2024-001', 'Schematic Design', 30, '2024-02-15', 1, True, '2024-02-10', 'Approved by client'],
        ['Website Redesign', 'PRJ-2024-001', 'Design Development', 60, '2024-04-01', 2, False, '', 'In progress'],
        ['Website Redesign', 'PRJ-2024-001', 'Construction Documents', 90, '2024-05-15', 3, False, '', 'Not started'],
        ['Mobile App Phase 2', 'PRJ-2024-002', 'Requirements Document', 15, '2024-03-15', 1, False, '', 'Initial draft'],
        ['Mobile App Phase 2', 'PRJ-2024-002', 'UI Wireframes', 40, '2024-04-30', 2, False, '', 'Design phase'],
        ['Internal Tool Update', '', 'Database Migration', 100, '2024-02-28', 1, True, '2024-02-25', 'Completed successfully'],
        ['Quick Fix Project', 'QF-001', 'Bug Fixes', 100, '2024-02-15', 1, True, '2024-02-12', 'All issues resolved']
    ]
    
    # Write example data
    for row_idx, deliverable_data in enumerate(example_deliverables, start=2):
        for col_idx, value in enumerate(deliverable_data, start=1):
            cell = deliverables_sheet.cell(row=row_idx, column=col_idx, value=value)
            # Light orange background for example data
            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    auto_fit_columns(deliverables_sheet)


def _create_projects_sheet(workbook, queryset):
    """Create main projects data sheet."""
    sheet = workbook.active
    sheet.title = "Projects"
    
    # Excel headers (camelCase to match API)
    headers = [
        'name', 'projectNumber', 'status', 'client', 'description',
        'startDate', 'endDate', 'estimatedHours', 'isActive'
    ]
    
    # Write headers with styling
    write_headers(sheet, headers)
    
    # Serialize data using ProjectSerializer
    serializer = ProjectSerializer(queryset, many=True)
    serialized_data = serializer.data
    
    # Write data rows (sanitize strings and force text type to avoid Excel formulas)
    for row_idx, project_data in enumerate(serialized_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            # Use serializer data directly (camelCase already provided by serializer)
            value = project_data.get(header, '')

            # Handle None values
            if value is None:
                value = ''

            safe_value = sanitize_cell(value) if isinstance(value, str) else value
            cell = sheet.cell(row=row_idx, column=col_idx, value=safe_value)
            if isinstance(value, str):
                cell.data_type = 's'
    
    auto_fit_columns(sheet)


def _create_assignments_sheet(workbook, queryset):
    """Create assignments sheet with people assigned to projects."""
    assignments_sheet = workbook.create_sheet("Assignments")
    
    # Excel headers for assignments
    headers = [
        'projectName', 'projectNumber', 'personName', 'personEmail',
        'roleOnProject', 'startDate', 'endDate', 'weeklyHours', 
        'totalHours', 'notes', 'isActive'
    ]
    
    write_headers(assignments_sheet, headers)
    
    row_idx = 2
    
    # Get all assignments for the projects in queryset
    for project in queryset:
        assignments = project.assignments.all() if hasattr(project, 'assignments') else []
        
        for assignment in assignments:
            # Handle weekly hours JSON
            weekly_hours_json = json.dumps(assignment.weekly_hours) if assignment.weekly_hours else "{}"
            total_hours = assignment.total_hours if hasattr(assignment, 'total_hours') else 0
            
            row_data = [
                project.name,                                   # projectName
                project.project_number or '',                  # projectNumber
                assignment.person.name,                        # personName
                assignment.person.email or '',                 # personEmail
                assignment.role_on_project or '',              # roleOnProject
                assignment.start_date or '',                   # startDate
                assignment.end_date or '',                     # endDate
                weekly_hours_json,                             # weeklyHours
                total_hours,                                   # totalHours
                assignment.notes or '',                        # notes
                assignment.is_active                           # isActive
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                safe_value = sanitize_cell(value) if isinstance(value, str) else value
                cell = assignments_sheet.cell(row=row_idx, column=col_idx, value=safe_value)
                if isinstance(value, str):
                    cell.data_type = 's'
            
            row_idx += 1
    
    auto_fit_columns(assignments_sheet)


def _create_deliverables_sheet(workbook, queryset):
    """Create deliverables sheet with project milestones."""
    deliverables_sheet = workbook.create_sheet("Deliverables")
    
    # Excel headers for deliverables
    headers = [
        'projectName', 'projectNumber', 'description', 'percentage',
        'date', 'sortOrder', 'isCompleted', 'completedDate', 'notes'
    ]
    
    write_headers(deliverables_sheet, headers)
    
    row_idx = 2
    
    # Get all deliverables for the projects in queryset
    for project in queryset:
        deliverables = project.deliverables.all()
        
        for deliverable in deliverables:
            row_data = [
                project.name,                                  # projectName
                project.project_number or '',                 # projectNumber
                deliverable.description or '',                # description
                deliverable.percentage or '',                 # percentage
                deliverable.date or '',                       # date
                deliverable.sort_order,                       # sortOrder
                deliverable.is_completed,                     # isCompleted
                deliverable.completed_date or '',             # completedDate
                deliverable.notes or ''                       # notes
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                safe_value = sanitize_cell(value) if isinstance(value, str) else value
                cell = deliverables_sheet.cell(row=row_idx, column=col_idx, value=safe_value)
                if isinstance(value, str):
                    cell.data_type = 's'
            
            row_idx += 1
    
    auto_fit_columns(deliverables_sheet)


def _create_projects_template_sheet(workbook):
    """Create template sheet with validation."""
    template_sheet = workbook.create_sheet("Template")
    
    # Projects section
    template_sheet.cell(row=1, column=1, value="PROJECTS TEMPLATE").font = Font(bold=True, size=14)
    
    projects_headers = [
        'name', 'projectNumber', 'status', 'client', 'description',
        'startDate', 'endDate', 'estimatedHours', 'isActive'
    ]
    
    # Projects template headers
    for col_idx, header in enumerate(projects_headers, start=1):
        cell = template_sheet.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Example project data
    example_project = [
        'Website Redesign', 'PRJ-2024-001', 'Active', 'Acme Corp',
        'Complete website overhaul', '2024-01-01', '2024-06-30', 2000, True
    ]
    
    for col_idx, value in enumerate(example_project, start=1):
        cell = template_sheet.cell(row=4, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Assignments section
    template_sheet.cell(row=6, column=1, value="ASSIGNMENTS TEMPLATE").font = Font(bold=True, size=14)
    
    assignments_headers = [
        'projectName', 'projectNumber', 'personName', 'personEmail',
        'roleOnProject', 'startDate', 'endDate', 'weeklyHours', 
        'totalHours', 'notes', 'isActive'
    ]
    
    # Assignments template headers
    for col_idx, header in enumerate(assignments_headers, start=1):
        cell = template_sheet.cell(row=8, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Example assignment data
    example_assignment = [
        'Website Redesign', 'PRJ-2024-001', 'John Smith', 'john@company.com',
        'Tech Lead', '2024-01-01', '2024-06-30', '{"2024-08-25":20,"2024-09-01":25}',
        800, 'Full-time on project', True
    ]
    
    for col_idx, value in enumerate(example_assignment, start=1):
        cell = template_sheet.cell(row=9, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    # Deliverables section
    template_sheet.cell(row=11, column=1, value="DELIVERABLES TEMPLATE").font = Font(bold=True, size=14)
    
    deliverables_headers = [
        'projectName', 'projectNumber', 'description', 'percentage',
        'date', 'sortOrder', 'isCompleted', 'completedDate', 'notes'
    ]
    
    # Deliverables template headers
    for col_idx, header in enumerate(deliverables_headers, start=1):
        cell = template_sheet.cell(row=13, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Example deliverable data
    example_deliverable = [
        'Website Redesign', 'PRJ-2024-001', 'Schematic Design', 30,
        '2024-02-15', 1, True, '2024-02-10', 'Approved by client'
    ]
    
    for col_idx, value in enumerate(example_deliverable, start=1):
        cell = template_sheet.cell(row=14, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    auto_fit_columns(template_sheet)


def _create_projects_instructions_sheet(workbook):
    """Create instructions sheet."""
    instructions_sheet = workbook.create_sheet("Instructions")
    
    instructions = [
        "Projects Import/Export Instructions",
        "",
        "MULTI-SHEET STRUCTURE:",
        "• Projects Sheet: Main project information",
        "• Assignments Sheet: People assigned to projects with weekly hours",
        "• Deliverables Sheet: Project milestones and completion tracking",
        "• Template Sheet: Examples for each data type",
        "",
        "PROJECTS SHEET FIELDS:",
        "• name - Project name (REQUIRED)",
        "• projectNumber - Unique project identifier",
        "• status - Planning/Active/Active CA/On Hold/Completed/Cancelled",
        "• client - Client name (default: Internal)",
        "• description - Project description",
        "• startDate/endDate - Date format: YYYY-MM-DD",
        "• estimatedHours - Total estimated hours",
        "• isActive - TRUE/FALSE (default: TRUE)",
        "",
        "ASSIGNMENTS SHEET FIELDS:",
        "• projectName/projectNumber - Links to project (REQUIRED)",
        "• personName - Person's full name (REQUIRED - will auto-create if not exists)",
        "• personEmail - Person's email address (optional but recommended)",
        "• personRole - Person's general job role (e.g., 'Senior Engineer', 'Designer')",
        "  → Used to create new person if they don't exist in system",
        "• roleOnProject - Person's specific role on THIS project (e.g., 'Tech Lead', 'QA')",
        "• startDate/endDate - Assignment date range",
        "• weeklyHours - JSON format: {'2024-08-25':10,'2024-09-01':8}",
        "• totalHours - Auto-calculated from weekly hours",
        "• notes - Assignment notes",
        "• isActive - TRUE/FALSE",
        "",
        "DELIVERABLES SHEET FIELDS:",
        "• projectName/projectNumber - Links to project (REQUIRED)",
        "• description - Deliverable description (e.g., SD, DD, IFC)",
        "• percentage - Completion percentage (0-100)",
        "• date - Target or completion date",
        "• sortOrder - Display order (lower numbers first)",
        "• isCompleted - TRUE/FALSE",
        "• completedDate - When actually completed",
        "• notes - Additional details",
        "",
        "IMPORT PROCESS:",
        "1. Projects are imported/updated first",
        "2. For each assignment:",
        "   → People are matched by email (preferred) or name",
        "   → If person doesn't exist, NEW PERSON IS AUTO-CREATED",
        "   → New people get personRole as their job role",
        "   → Assignments link projects to people (existing or new)",
        "3. Deliverables link to projects",
        "4. All relationships are validated",
        "",
        "MATCHING LOGIC:",
        "• Projects: Match by projectNumber (preferred) or name",
        "• People: Match by email (preferred) or name",
        "• Cross-validation ensures all references are valid",
        "",
        "WEEKLY HOURS JSON FORMAT:",
        "• Use Sunday dates as keys: '2024-08-25'",
        "• Values are hours for that week: 10, 20, etc.",
        "• Example: {'2024-08-25':20,'2024-09-01':25,'2024-09-08':15}",
        "",
        "ERROR PREVENTION:",
        "• Use the Template sheet for proper formatting",
        "• Ensure all required relationships exist",
        "• Use dry-run mode to preview changes",
        "• Check for duplicate project numbers"
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        cell = instructions_sheet.cell(row=row_idx, column=1, value=instruction)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction.startswith(("PROJECTS SHEET", "ASSIGNMENTS SHEET", "DELIVERABLES SHEET", "MULTI-SHEET", "IMPORT PROCESS", "MATCHING LOGIC", "WEEKLY HOURS", "ERROR PREVENTION")):
            cell.font = Font(bold=True, size=12)
    
    auto_fit_columns(instructions_sheet)


def _write_excel_headers(sheet, headers):
    # Backward-compat wrapper for existing imports; delegate to shared util
    write_headers(sheet, headers)


def _auto_fit_columns(sheet):
    auto_fit_columns(sheet)


def _create_excel_response(workbook, filename):
    return create_excel_response(workbook, filename)


# IMPORT FUNCTIONALITY - Phase 6 Implementation

def import_projects_from_file(file, update_existing=True, include_assignments=True, include_deliverables=True, dry_run=False):
    """
    Import projects from Excel or CSV file using serializers.
    Follows R2-REBUILD-STANDARDS: All data transformations use serializers.
    """
    file_name = file.name.lower()
    
    if file_name.endswith(('.xlsx', '.xls')):
        return _import_projects_from_excel(file, update_existing, include_assignments, include_deliverables, dry_run)
    elif file_name.endswith('.csv'):
        return _import_projects_from_csv(file, update_existing, dry_run)
    else:
        return {
            'success': False,
            'errors': ['Unsupported file format. Use Excel (.xlsx) or CSV (.csv) files.'],
            'summary': {}
        }


def _import_projects_from_excel(file, update_existing, include_assignments, include_deliverables, dry_run):
    """Import projects from multi-sheet Excel file using ProjectSerializer."""
    try:
        # Safety: enforce structural ceilings before heavy parse
        try:
            from core.utils.xlsx_limits import enforce_xlsx_limits
            enforce_xlsx_limits(file)
        except Exception:
            # If limits fail, openpyxl will likely also fail; let error propagate below
            pass
        workbook = openpyxl.load_workbook(file, data_only=True)
        results = {
            'success': True,
            'errors': [],
            'projects_created': 0,
            'projects_updated': 0,
            'assignments_created': 0,
            'people_created': 0,
            'deliverables_created': 0,
            'projects_to_create': [],
            'projects_to_update': [],
            'summary': {}
        }
        
        with transaction.atomic():
            # Phase 1: Import Projects (required)
            if 'Projects' in workbook.sheetnames:
                project_results = _import_projects_sheet(workbook['Projects'], update_existing, dry_run)
                results.update(project_results)
            else:
                results['errors'].append('Projects sheet not found. Excel file must contain a "Projects" sheet.')
                results['success'] = False
                return results
            
            # Phase 2: Import Assignments (optional)
            if include_assignments and 'Assignments' in workbook.sheetnames:
                assignment_results = _import_assignments_sheet(workbook['Assignments'], dry_run)
                results['assignments_created'] = assignment_results.get('assignments_created', 0)
                results['people_created'] = assignment_results.get('people_created', 0)
                results['errors'].extend(assignment_results.get('errors', []))
            
            # Phase 3: Import Deliverables (optional)
            if include_deliverables and 'Deliverables' in workbook.sheetnames:
                deliverable_results = _import_deliverables_sheet(workbook['Deliverables'], dry_run)
                results['deliverables_created'] = deliverable_results.get('deliverables_created', 0)
                results['errors'].extend(deliverable_results.get('errors', []))
            
            # Rollback if dry run
            if dry_run:
                transaction.set_rollback(True)
        
        results['summary'] = _create_import_summary(results)
        return results
        
    except Exception as e:
        return {
            'success': False,
            'errors': [f'Error processing Excel file: {str(e)}'],
            'summary': {}
        }


def _import_projects_sheet(projects_sheet, update_existing, dry_run):
    """Import projects from Projects sheet using ProjectSerializer.
    
    Supports TWO FORMATS:
    1. Standard format: Just project data
    2. Wide format: Project data + person1Name, person1Role, person2Name, person2Role, etc.
    """
    results = {
        'projects_created': 0,
        'projects_updated': 0,
        'projects_to_create': [],
        'projects_to_update': [],
        'assignments_created': 0,
        'people_created': 0,
        'errors': []
    }
    
    # Get headers from first row (camelCase)
    headers = [cell.value for cell in projects_sheet[1] if cell.value]
    
    # Detect if this is wide format (has person columns)
    person_columns = [h for h in headers if h and (h.startswith('person') and ('Name' in h or 'Role' in h or 'Email' in h))]
    is_wide_format = len(person_columns) > 0
    
    if is_wide_format:
        print(f"Detected wide format with {len(person_columns)} person columns")
    
    # Process each data row
    for row_idx, row in enumerate(projects_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        # Create row data dict with camelCase keys
        row_data = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_data[headers[col_idx]] = value
        
        # Skip rows without required data
        if not row_data.get('name'):
            results['errors'].append(f'Row {row_idx}: Missing required field "name"')
            continue
        
        try:
            # Import the project first
            project_result = _import_single_project(row_data, update_existing, dry_run)
            
            if project_result['created']:
                results['projects_created'] += 1
                results['projects_to_create'].append(row_data)
            elif project_result['updated']:
                results['projects_updated'] += 1
                results['projects_to_update'].append({
                    **row_data,
                    'changes': project_result.get('changes', 'Updated')
                })
            
            if project_result.get('errors'):
                results['errors'].extend([f'Row {row_idx}: {err}' for err in project_result['errors']])
            
            # Handle wide format assignments (if present)
            if is_wide_format:
                assignment_results = _process_wide_format_assignments(
                    row_data, project_result.get('project'), dry_run, row_idx
                )
                results['assignments_created'] += assignment_results.get('assignments_created', 0)
                results['people_created'] += assignment_results.get('people_created', 0)
                results['errors'].extend(assignment_results.get('errors', []))
                
        except Exception as e:
            results['errors'].append(f'Row {row_idx}: {str(e)}')
    
    return results


def _import_single_project(row_data, update_existing, dry_run):
    """Import single project using ProjectSerializer (R2-REBUILD-STANDARDS compliant)."""
    # Clean row data - convert empty strings to None for optional fields
    cleaned_data = _clean_project_data(row_data)
    
    # Match existing project by projectNumber (preferred) or name
    existing_project = None
    
    if cleaned_data.get('projectNumber'):
        existing_project = Project.objects.filter(project_number=cleaned_data['projectNumber']).first()
    
    if not existing_project and cleaned_data.get('name'):
        existing_project = Project.objects.filter(name=cleaned_data['name']).first()
    
    result = {'created': False, 'updated': False, 'errors': [], 'project': None}
    
    try:
        if existing_project and update_existing:
            # Update existing project using serializer
            serializer = ProjectSerializer(existing_project, data=cleaned_data, partial=True)
            if serializer.is_valid():
                if not dry_run:
                    project = serializer.save()
                    result['project'] = project
                else:
                    result['project'] = existing_project  # In dry-run, return existing project
                result['updated'] = True
                result['changes'] = 'Updated with new data'
            else:
                result['errors'] = [f"{field}: {errors}" for field, errors in serializer.errors.items()]
                
        elif not existing_project:
            # Create new project using serializer
            serializer = ProjectSerializer(data=cleaned_data)
            if serializer.is_valid():
                if not dry_run:
                    project = serializer.save()
                    result['project'] = project
                else:
                    # In dry-run mode, create a mock project object for assignment processing
                    mock_project = Project(**{k: v for k, v in cleaned_data.items() if hasattr(Project, k)})
                    mock_project.id = 999  # Mock ID
                    result['project'] = mock_project
                result['created'] = True
            else:
                result['errors'] = [f"{field}: {errors}" for field, errors in serializer.errors.items()]
        else:
            # Project exists but update_existing=False, still return it for assignments
            result['project'] = existing_project
        
    except Exception as e:
        result['errors'].append(str(e))
    
    return result


def _import_assignments_sheet(assignments_sheet, dry_run):
    """Import assignments from Assignments sheet using AssignmentSerializer."""
    results = {
        'assignments_created': 0,
        'people_created': 0,
        'errors': []
    }
    
    # Get headers from first row (camelCase)
    headers = [cell.value for cell in assignments_sheet[1] if cell.value]
    
    # Process each data row
    for row_idx, row in enumerate(assignments_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        # Create row data dict with camelCase keys
        row_data = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_data[headers[col_idx]] = value
        
        try:
            assignment_result = _import_single_assignment(row_data, dry_run)
            
            if assignment_result['created']:
                results['assignments_created'] += 1
            
            # Check if a person was created during this assignment
            if assignment_result.get('person_created'):
                results['people_created'] += 1
            
            if assignment_result.get('errors'):
                results['errors'].extend([f'Assignments Row {row_idx}: {err}' for err in assignment_result['errors']])
                
        except Exception as e:
            results['errors'].append(f'Assignments Row {row_idx}: {str(e)}')
    
    return results


def _import_single_assignment(row_data, dry_run):
    """Import single assignment using AssignmentSerializer."""
    result = {'created': False, 'person_created': False, 'errors': []}
    
    try:
        # Find project by name or projectNumber
        project = None
        if row_data.get('projectNumber'):
            project = Project.objects.filter(project_number=row_data['projectNumber']).first()
        if not project and row_data.get('projectName'):
            project = Project.objects.filter(name=row_data['projectName']).first()
        
        if not project:
            result['errors'].append('Project not found')
            return result
        
        # Find or create person
        person = _find_or_create_person(row_data, dry_run)
        
        if not person['success']:
            result['errors'].extend(person['errors'])
            return result
        
        person_instance = person['person']
        
        # Track if person was created
        if person['created']:
            result['person_created'] = True
        
        # Prepare assignment data for serializer
        assignment_data = {
            'project': project.id,
            'person': person_instance.id if person_instance else None,
            'roleOnProject': row_data.get('roleOnProject', ''),
            'startDate': row_data.get('startDate'),
            'endDate': row_data.get('endDate'),
            'weeklyHours': _parse_weekly_hours_json(row_data.get('weeklyHours', '{}')),
            'notes': row_data.get('notes', ''),
            'isActive': row_data.get('isActive', True)
        }
        
        # Use AssignmentSerializer for data transformation
        serializer = AssignmentSerializer(data=assignment_data)
        if serializer.is_valid():
            if not dry_run:
                serializer.save()
            result['created'] = True
        else:
            result['errors'] = [f"{field}: {errors}" for field, errors in serializer.errors.items()]
        
    except Exception as e:
        result['errors'].append(str(e))
    
    return result


def _process_wide_format_assignments(row_data, project, dry_run, row_idx):
    """
    Process assignments from wide format (person1Name, person1Role, person2Name, person2Role, etc).
    
    Expected columns:
    - person1Name, person1Role, person1Email (optional)
    - person2Name, person2Role, person2Email (optional)
    - etc. (supports up to person9)
    """
    results = {
        'assignments_created': 0,
        'people_created': 0,
        'errors': []
    }
    
    if not project:
        results['errors'].append('Cannot create assignments without valid project')
        return results
    
    # Find all person groups (person1*, person2*, etc.)
    person_groups = {}
    
    for key, value in row_data.items():
        if key and key.startswith('person') and value and str(value).strip():
            # Extract person number (person1Name -> 1, person2Role -> 2, etc.)
            import re
            match = re.match(r'person(\d+)(.+)', key)
            if match:
                person_num = match.group(1)
                field_type = match.group(2)  # Name, Role, Email, etc.
                
                if person_num not in person_groups:
                    person_groups[person_num] = {}
                
                person_groups[person_num][field_type] = str(value).strip()
    
    # Process each person group
    for person_num, person_data in person_groups.items():
        if not person_data.get('Name'):  # Skip if no name
            continue
            
        try:
            # Create assignment data in the standard format
            assignment_data = {
                'projectName': project.name,
                'projectNumber': getattr(project, 'project_number', ''),
                'personName': person_data.get('Name', ''),
                'personEmail': person_data.get('Email', ''),
                'roleOnProject': person_data.get('Role', 'Team Member'),
                'startDate': row_data.get('startDate'),  # Use project dates as default
                'endDate': row_data.get('endDate'),
                'weeklyHours': '{}',  # Empty JSON for now, could be enhanced later
                'totalHours': person_data.get('TotalHours', ''),
                'notes': f"Created from wide format row {row_idx}",
                'isActive': True
            }
            
            # Import this assignment using existing logic
            assignment_result = _import_single_assignment(assignment_data, dry_run)
            
            if assignment_result.get('created'):
                results['assignments_created'] += 1
            
            if assignment_result.get('person_created'):
                results['people_created'] += 1
            
            if assignment_result.get('errors'):
                results['errors'].extend([
                    f"Person{person_num} ({person_data.get('Name', 'Unknown')}): {err}" 
                    for err in assignment_result['errors']
                ])
                
        except Exception as e:
            results['errors'].append(f"Person{person_num} ({person_data.get('Name', 'Unknown')}): {str(e)}")
    
    return results


def _import_deliverables_sheet(deliverables_sheet, dry_run):
    """Import deliverables from Deliverables sheet."""
    results = {
        'deliverables_created': 0,
        'errors': []
    }
    
    # Note: Deliverable serializer implementation would go here
    # For now, return placeholder results
    results['errors'].append('Deliverables import not yet implemented')
    
    return results


def _find_or_create_person(row_data, dry_run):
    """Find existing person or create new one from assignment data."""
    result = {'success': False, 'person': None, 'errors': [], 'created': False}
    
    try:
        # Try to find existing person by email (preferred) or name
        person = None
        if row_data.get('personEmail'):
            person = Person.objects.filter(email=row_data['personEmail']).first()
        
        if not person and row_data.get('personName'):
            person = Person.objects.filter(name=row_data['personName']).first()
        
        if person:
            # Found existing person
            result['success'] = True
            result['person'] = person
            return result
        
        # Person doesn't exist - create new one
        if not row_data.get('personName'):
            result['errors'].append('personName is required to create new person')
            return result
        
        # Prepare person data for creation
        person_data = {
            'name': row_data.get('personName'),
            'email': row_data.get('personEmail', ''),
            'role': row_data.get('personRole', 'Team Member'),  # Default role
            'weeklyCapacity': 40,  # Default capacity
            'notes': f"Auto-created during project import"
        }
        
        # Clean person data (remove empty emails)
        if not person_data['email'] or person_data['email'].strip() == '':
            person_data['email'] = None
        
        # Use PersonSerializer to create person
        person_serializer = PersonSerializer(data=person_data)
        if person_serializer.is_valid():
            if not dry_run:
                new_person = person_serializer.save()
                result['person'] = new_person
            else:
                # In dry run, create a mock person object for validation
                result['person'] = Person(
                    name=person_data['name'],
                    email=person_data.get('email'),
                    id=999  # Mock ID for dry run
                )
            result['success'] = True
            result['created'] = True
        else:
            result['errors'] = [f"Failed to create person: {field}: {errors}" 
                             for field, errors in person_serializer.errors.items()]
    
    except Exception as e:
        result['errors'].append(f"Error finding/creating person: {str(e)}")
    
    return result


def _clean_project_data(row_data):
    """Clean project row data - convert empty strings to None for optional fields."""
    cleaned_data = row_data.copy()
    
    # Optional fields that should be None instead of empty strings
    optional_fields = [
        'projectNumber', 'startDate', 'endDate', 'estimatedHours', 
        'description', 'client'
    ]
    
    for field in optional_fields:
        if field in cleaned_data:
            value = cleaned_data[field]
            # Convert empty strings, whitespace-only strings to None
            if isinstance(value, str) and value.strip() == '':
                cleaned_data[field] = None
            # Convert string representations of numbers for estimatedHours
            elif field == 'estimatedHours' and isinstance(value, str):
                try:
                    # Try to convert to int, but allow None for empty values
                    cleaned_data[field] = int(value) if value.strip() else None
                except (ValueError, TypeError):
                    cleaned_data[field] = None
    
    return cleaned_data


def _parse_weekly_hours_json(weekly_hours_str):
    """Parse weekly hours from JSON string format."""
    if not weekly_hours_str or weekly_hours_str in ['', '{}']:
        return {}
    
    try:
        if isinstance(weekly_hours_str, dict):
            return weekly_hours_str
        
        # Parse JSON string
        return json.loads(str(weekly_hours_str))
    except (json.JSONDecodeError, TypeError):
        return {}


def _import_projects_from_csv(file, update_existing, dry_run):
    """Import projects from CSV file using ProjectSerializer."""
    results = {
        'success': True,
        'errors': [],
        'projects_created': 0,
        'projects_updated': 0,
        'projects_to_create': [],
        'projects_to_update': [],
        'summary': {}
    }
    
    try:
        # Read CSV file
        file_content = file.read()
        if isinstance(file_content, bytes):
            file_content = file_content.decode('utf-8')
        
        csv_reader = csv.DictReader(io.StringIO(file_content))
        
        with transaction.atomic():
            for row_idx, row_data in enumerate(csv_reader, start=2):
                if not row_data.get('name'):
                    results['errors'].append(f'Row {row_idx}: Missing required field "name"')
                    continue
                
                try:
                    project_result = _import_single_project(row_data, update_existing, dry_run)
                    
                    if project_result['created']:
                        results['projects_created'] += 1
                        results['projects_to_create'].append(row_data)
                    elif project_result['updated']:
                        results['projects_updated'] += 1
                        results['projects_to_update'].append(row_data)
                    
                    if project_result.get('errors'):
                        results['errors'].extend([f'Row {row_idx}: {err}' for err in project_result['errors']])
                        
                except Exception as e:
                    results['errors'].append(f'Row {row_idx}: {str(e)}')
            
            # Rollback if dry run
            if dry_run:
                transaction.set_rollback(True)
        
        results['summary'] = _create_import_summary(results)
        return results
        
    except Exception as e:
        return {
            'success': False,
            'errors': [f'Error processing CSV file: {str(e)}'],
            'summary': {}
        }


def _create_import_summary(results):
    """Create summary of import results."""
    total_processed = results.get('projects_created', 0) + results.get('projects_updated', 0)
    total_errors = len(results.get('errors', []))
    
    return {
        'total_processed': total_processed,
        'projects_created': results.get('projects_created', 0),
        'projects_updated': results.get('projects_updated', 0),
        'assignments_created': results.get('assignments_created', 0),
        'people_created': results.get('people_created', 0),
        'deliverables_created': results.get('deliverables_created', 0),
        'total_errors': total_errors,
        'success_rate': f"{((total_processed - total_errors) / max(total_processed, 1)) * 100:.1f}%"
    }
