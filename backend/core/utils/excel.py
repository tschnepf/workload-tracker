"""
Excel utility helpers shared across apps.

Keep pure (no Django model imports) to avoid cycles.
"""
from __future__ import annotations

from typing import Iterable
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def write_headers(sheet, headers: Iterable[str]) -> None:
    """Write a header row with consistent styling."""
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def auto_fit_columns(sheet, max_width: int = 50, padding: int = 2) -> None:
    """Auto-fit column widths up to a maximum width."""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        width = min(max_length + padding, max_width)
        sheet.column_dimensions[column_letter].width = width


def create_excel_response(workbook, filename: str) -> HttpResponse:
    """Return an HttpResponse with the given workbook serialized as XLSX."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

