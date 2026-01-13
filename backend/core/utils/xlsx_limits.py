import os
from typing import Union


def _open_workbook_for_limits(path_or_file):
    import openpyxl
    # Prefer path for streaming-friendly open; fall back to file-like
    if isinstance(path_or_file, (str, bytes, os.PathLike)):
        return openpyxl.load_workbook(path_or_file, read_only=True, data_only=True)
    return openpyxl.load_workbook(path_or_file, read_only=True, data_only=True)


def enforce_xlsx_limits(
    path_or_file: Union[str, os.PathLike, object],
    max_sheets: int = 10,
    max_rows_per_sheet: int = 100_000,
    max_total_cells: int = 5_000_000,
) -> None:
    """Raise ValueError if workbook exceeds safety ceilings.

    Limits inspired by common guardrails against resource-exhaustion (zip bombs).
    Evaluates sheet count, per-sheet rows, and total approximate cells (rows*cols).
    """
    try:
        wb = _open_workbook_for_limits(path_or_file)
    except Exception as e:
        raise ValueError(f"invalid_excel: {e.__class__.__name__}")

    try:
        sheets = wb.sheetnames
        if len(sheets) > max_sheets:
            raise ValueError("xlsx_sheets_exceeded")

        total_cells = 0
        for name in sheets:
            ws = wb[name]
            rows = int(getattr(ws, 'max_row', 0) or 0)
            cols = int(getattr(ws, 'max_column', 0) or 0)
            if rows > max_rows_per_sheet:
                raise ValueError("xlsx_rows_exceeded")
            # Accumulate approximate cells; cap if obviously exploding
            total_cells += rows * max(cols, 0)
            if total_cells > max_total_cells:
                raise ValueError("xlsx_total_cells_exceeded")
    finally:
        try:
            wb.close()
        except Exception:  # nosec B110
            pass

