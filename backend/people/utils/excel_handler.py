"""
Excel handler for People import/export functionality.
Follows R2-REBUILD-STANDARDS: snake_case Python, camelCase Excel headers.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from ..serializers import PersonSerializer
from ..models import Person
from core.utils.excel import write_headers, auto_fit_columns, create_excel_response


def export_people_to_excel(queryset, filename=None):
    """Export people queryset to Excel with multiple sheets."""
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"people_export_{timestamp}.xlsx"
    
    workbook = openpyxl.Workbook()
    
    # Create People sheet
    _create_people_sheet(workbook, queryset)
    
    # Create Template sheet
    _create_people_template_sheet(workbook)
    
    # Create Instructions sheet
    _create_people_instructions_sheet(workbook)
    
    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    return create_excel_response(workbook, filename)


def _create_people_sheet(workbook, queryset):
    """Create main people data sheet."""
    sheet = workbook.active
    sheet.title = "People"
    
    # Excel headers (camelCase to match API) - ID first for primary lookup
    headers = [
        'id', 'name', 'role', 'email', 'phone', 'location', 
        'weeklyCapacity', 'departmentName', 'hireDate', 
        'notes', 'isActive'
    ]
    
    # Write headers with styling
    write_headers(sheet, headers)
    
    # Serialize data using PersonSerializer
    serializer = PersonSerializer(queryset, many=True)
    serialized_data = serializer.data
    
    # Write data rows
    for row_idx, person_data in enumerate(serialized_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = person_data.get(header, '')
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    auto_fit_columns(sheet)


def _create_people_template_sheet(workbook):
    """Create template sheet with validation."""
    template_sheet = workbook.create_sheet("Template")
    
    headers = [
        'id', 'name', 'role', 'email', 'phone', 'location',
        'weeklyCapacity', 'departmentName', 'hireDate', 
        'notes', 'isActive'
    ]
    
    write_headers(template_sheet, headers)
    
    # Add example row (leave ID blank for new people, or use existing ID for updates)
    example_data = [
        '', 'John Smith', 'Senior Engineer', 'john@company.com',
        '555-0123', 'New York', 40, 'Engineering',
        '2023-01-15', 'Team lead', True
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = template_sheet.cell(row=2, column=col_idx, value=value)
        cell.fill = PatternFill(start_color="E6F3FF", 
                               end_color="E6F3FF", 
                               fill_type="solid")
    
    auto_fit_columns(template_sheet)


def _create_people_instructions_sheet(workbook):
    """Create instructions sheet."""
    instructions_sheet = workbook.create_sheet("Instructions")
    
    instructions = [
        "People Import Instructions - ID-Based Import/Export",
        "",
        "Primary Lookup Logic:",
        "• id - Primary lookup field (blank for new people)",
        "• If ID exists, updates that specific person",
        "• If ID is blank, falls back to email then name lookup",
        "• If no match found, creates new person",
        "",
        "Required Fields:",
        "• name - Person's full name",
        "",
        "Optional Fields:",
        "• id - Person ID (leave blank for new people)",
        "• role - Job title (default: Engineer)",
        "• email - Contact email address", 
        "• phone - Phone number",
        "• location - Work location",
        "• weeklyCapacity - Hours per week (default: 36)",
        "• departmentName - Department name",
        "• hireDate - Start date (YYYY-MM-DD format)",
        "• notes - Additional notes",
        "• isActive - TRUE/FALSE (default: TRUE)",
        "",
        "Batch Name Editing Workflow:",
        "1. Export people to Excel (includes IDs)",
        "2. Edit names in Excel (keep IDs intact)",
        "3. Import back - names will be updated using ID lookup",
        "",
        "Import Process:",
        "1. Fill out the Template sheet",
        "2. Save as Excel file",
        "3. Use Django Admin import function",
        "4. Review validation results"
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        cell = instructions_sheet.cell(row=row_idx, column=1, value=instruction)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)


def _write_excel_headers(sheet, headers):
    # Backward-compat wrapper for existing imports; delegate to shared util
    write_headers(sheet, headers)


def _auto_fit_columns(sheet):
    auto_fit_columns(sheet)


def _create_excel_response(workbook, filename):
    return create_excel_response(workbook, filename)


def import_people_from_excel(file, update_existing=True, dry_run=False, progress_callback=None):
    """Import people from Excel file."""
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        if not headers:
            return {
                'success': False,
                'error': 'No headers found in Excel file',
                'total_rows': 0,
                'success_count': 0,
                'error_count': 1,
                'errors': ['No headers found in Excel file']
            }
        
        results = {
            'success': True,
            'total_rows': 0,
            'success_count': 0,
            'updated_count': 0,
            'error_count': 0,
            'errors': [],
            'warnings': [],
            'success_items': [],
            'dry_run': dry_run
        }
        
        # Get total rows for progress calculation
        all_rows = list(sheet.iter_rows(min_row=2))
        total_data_rows = sum(1 for row in all_rows if any(cell.value for cell in row))
        
        if progress_callback:
            progress_callback({
                'stage': 'starting',
                'message': f'Starting import of {total_data_rows} people...',
                'progress': 0,
                'total': total_data_rows
            })
        
        # Process data rows with progress tracking
        for row_idx, row in enumerate(all_rows):
            if not any(cell.value for cell in row):
                continue  # Skip empty rows
                
            results['total_rows'] += 1
            row_num = row_idx + 2  # Actual row number in Excel
            row_data = {}
            
            # Map row values to headers
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    cell_value = row[col_idx].value
                    if cell_value is not None:
                        row_data[header] = cell_value
            
            # Process individual row
            row_result = _process_people_row(row_data, update_existing, dry_run, row_num)
            
            if row_result['success']:
                results['success_count'] += 1
                if row_result.get('updated'):
                    results['updated_count'] += 1
                results['success_items'].append(row_result['message'])
            else:
                results['error_count'] += 1
                results['errors'].append(f"Row {row_num}: {row_result['error']}")
            
            # Report progress every 10 records or at the end
            if progress_callback and (results['total_rows'] % 10 == 0 or results['total_rows'] == total_data_rows):
                progress_percent = int((results['total_rows'] / total_data_rows) * 100)
                progress_callback({
                    'stage': 'processing',
                    'message': f'Processed {results["total_rows"]}/{total_data_rows} people (Success: {results["success_count"]}, Errors: {results["error_count"]})',
                    'progress': progress_percent,
                    'total': total_data_rows
                })
        
        if progress_callback:
            progress_callback({
                'stage': 'complete',
                'message': f'Import completed: {results["success_count"]} successful, {results["error_count"]} errors',
                'progress': 100,
                'total': total_data_rows
            })
        
        return results
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to process Excel file: {str(e)}',
            'total_rows': 0,
            'success_count': 0,
            'error_count': 1,
            'errors': [f'Failed to process Excel file: {str(e)}']
        }


def _process_people_row(row_data, update_existing, dry_run, row_num):
    """Process single person row using PersonSerializer with ID-based lookup."""
    try:
        # Check if person exists using ID-based lookup hierarchy
        existing_person = None
        lookup_method = 'none'
        
        # First priority: Try to find by ID (most reliable)
        if 'id' in row_data and row_data['id']:
            try:
                person_id = int(row_data['id'])
                existing_person = Person.objects.get(id=person_id)
                lookup_method = 'id'
            except (ValueError, Person.DoesNotExist):
                return {
                    'success': False,
                    'error': f"Person with ID {row_data['id']} not found. ID may be invalid or person may have been deleted."
                }
        
        # Second priority: Try to find by email (if ID not provided)
        elif 'email' in row_data and row_data['email']:
            try:
                existing_person = Person.objects.get(email=row_data['email'])
                lookup_method = 'email'
            except Person.DoesNotExist:
                pass
        
        # Third priority: Try to find by name (least reliable)
        if not existing_person and 'name' in row_data and row_data['name']:
            try:
                existing_person = Person.objects.get(name=row_data['name'])
                lookup_method = 'name'
            except Person.DoesNotExist:
                pass
            except Person.MultipleObjectsReturned:
                # If multiple people have the same name, we can't safely update
                return {
                    'success': False,
                    'error': f"Multiple people found with name '{row_data['name']}'. Please use ID or unique email for updates."
                }
        
        # Update existing person
        if existing_person:
            if not update_existing:
                identifier = f"ID:{existing_person.id}" if lookup_method == 'id' else row_data.get('email', row_data.get('name', 'Unknown'))
                return {
                    'success': True,
                    'updated': False,
                    'message': f"Skipped existing person: {row_data.get('name', 'Unknown')} (found by {lookup_method}: {identifier})"
                }
            
            if not dry_run:
                # Remove 'id' from row_data to prevent serializer conflicts
                update_data = {k: v for k, v in row_data.items() if k != 'id'}
                serializer = PersonSerializer(existing_person, data=update_data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    identifier = f"ID:{existing_person.id}" if lookup_method == 'id' else row_data.get('email', row_data.get('name', 'Unknown'))
                    return {
                        'success': True,
                        'updated': True,
                        'message': f"Updated: {row_data.get('name', 'Unknown')} (found by {lookup_method}: {identifier})"
                    }
                else:
                    return {
                        'success': False,
                        'error': f"Validation errors: {serializer.errors}"
                    }
            else:
                identifier = f"ID:{existing_person.id}" if lookup_method == 'id' else row_data.get('email', row_data.get('name', 'Unknown'))
                return {
                    'success': True,
                    'updated': True,
                    'message': f"[DRY RUN] Would update: {row_data.get('name', 'Unknown')} (found by {lookup_method}: {identifier})"
                }
        
        # Create new person
        else:
            if not dry_run:
                # Remove 'id' from row_data for new person creation
                create_data = {k: v for k, v in row_data.items() if k != 'id'}
                serializer = PersonSerializer(data=create_data)
                if serializer.is_valid():
                    new_person = serializer.save()
                    identifier = row_data.get('email', row_data.get('name', 'Unknown'))
                    return {
                        'success': True,
                        'updated': False,
                        'message': f"Created: {row_data.get('name', 'Unknown')} (new ID: {new_person.id}) ({identifier})"
                    }
                else:
                    return {
                        'success': False,
                        'error': f"Validation errors: {serializer.errors}"
                    }
            else:
                identifier = row_data.get('email', row_data.get('name', 'Unknown'))
                return {
                    'success': True,
                    'updated': False,
                    'message': f"[DRY RUN] Would create: {row_data.get('name', 'Unknown')} ({identifier})"
                }
                
    except Exception as e:
        return {
            'success': False,
            'error': f"Unexpected error: {str(e)}"
        }
